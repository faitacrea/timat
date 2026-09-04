# -*- coding: utf-8 -*-
"""Kit de gestion TiMat pour assistante maternelle.

Quatre contrats dans un seul classeur. Le premier jet en tenait un seul, ce
qui obligeait une assmat avec quatre enfants à jongler entre quatre fichiers
sans jamais voir son mois en entier — exactement le problème que le kit est
censé résoudre.

Repères de contrôle, à vérifier en ouvrant le fichier :
  — la ligne d'exemple va de 07:30 à 18:00, soit 10,50 heures, et son
    indemnité d'entretien doit afficher 4,57 € (10,50 × 0,435) ;
  — une journée plus courte, de 07:30 à 12:00, doit afficher le plancher de
    2,65 € et non 1,96 €. C'est le comportement le moins évident du classeur.

Tout est formule : la lectrice change un paramètre de contrat et l'année
entière se recalcule. Rien n'est calculé en Python puis figé dans une cellule.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

MARINE = "2E4859"; TERRA = "C84B31"; LIGNE = "E4DCD0"; SAUGE = "5DA9A1"
F = "Arial"
h1 = Font(name=F, size=16, bold=True, color="FFFFFF")
h2 = Font(name=F, size=11, bold=True, color="FFFFFF")
gras = Font(name=F, size=10, bold=True, color=MARINE)
norm = Font(name=F, size=10, color=MARINE)
saisie = Font(name=F, size=10, bold=True, color="0000FF")
petit = Font(name=F, size=9, color="6B7A82")
fond_titre = PatternFill("solid", fgColor=MARINE)
fond_entete = PatternFill("solid", fgColor=TERRA)
fond_saisie = PatternFill("solid", fgColor="FFF9E6")
fond_doux = PatternFill("solid", fgColor="F4F1EA")
bord = Border(*[Side(style="thin", color=LIGNE)] * 4)
EUR = '#,##0.00 "€"'; HEU = '#,##0.00'; DATE = "DD/MM/YYYY"; HM = "HH:MM"; ENT = "#,##0"

# Quatre contrats : au-delà, l'agrément lui-même devient l'exception.
NB_ENFANTS = 4
COLS_ENF = ["C", "D", "E", "F"]
PREMIERE, DERNIERE = 4, 1003          # lignes de saisie de l'onglet Heures
LIG = {                                # lignes de l'onglet Contrats
    "prenom": 5, "debut": 6, "taux": 7, "heuresSem": 8, "semaines": 9,
    "entretienH": 10, "plancher": 11, "repas": 12, "km": 13, "semCongés": 14,
    "heuresMens": 16, "brutMens": 17, "ctrlMin": 18, "ctrlCMG": 19,
}

wb = Workbook()

# ------------------------------------------------------------- Mode d'emploi
ws = wb.active; ws.title = "Mode d'emploi"
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 104
ws["B2"] = "Kit de gestion — assistante maternelle"
ws["B2"].font = Font(name=F, size=18, bold=True, color=MARINE)
ws["B3"] = "TiMat · timat.app"; ws["B3"].font = petit

lignes = [
    ("", ""),
    ("titre", "Ce que fait ce classeur"),
    ("p", "Vous saisissez vos horaires jour par jour, pour jusqu'à quatre enfants. Le classeur calcule les heures, "
          "les heures complémentaires, l'indemnité d'entretien, les repas et les kilomètres — mois par mois pour "
          "l'enfant que vous choisissez, puis pour l'année entière et pour tous les enfants ensemble."),
    ("", ""),
    ("titre", "Les trois règles à retenir"),
    ("p", "1. Les cellules sur fond crème et en bleu sont les seules à remplir. Tout le reste est calculé : "
          "si vous écrivez par-dessus une formule, elle est perdue."),
    ("p", "2. Commencez par l'onglet « Contrats ». Rien ne se calcule tant qu'il est vide."),
    ("p", "3. Chaque journée saisie porte un numéro d'enfant, de 1 à 4. C'est ce numéro qui relie la journée à son "
          "contrat : sans lui, la ligne ne compte nulle part."),
    ("", ""),
    ("titre", "Les onglets, dans l'ordre"),
    ("p", "Contrats — un enfant par colonne : taux horaire, heures, indemnités. À remplir en premier."),
    ("p", "Heures — une ligne par journée d'accueil, avec le numéro de l'enfant."),
    ("p", "Mois — le détail mensuel de l'enfant que vous choisissez en haut de l'onglet."),
    ("p", "Année — les totaux annuels des quatre enfants côte à côte, et les congés payés."),
    ("", ""),
    ("titre", "Ce que ce classeur ne fait pas"),
    ("p", "Il ne remplace pas un bulletin de salaire et ne calcule pas les cotisations : c'est Pajemploi qui les "
          "établit à partir de ce que vous déclarez. Il vous donne les montants à déclarer, et de quoi les vérifier."),
    ("p", "Il ne calcule pas la majoration des heures au-delà de 45 heures par semaine : ce taux se négocie au "
          "contrat, et lui seul fait foi."),
    ("", ""),
    ("titre", "Les repères de la convention collective"),
    ("p", "Minimum conventionnel : 4,20 € brut de l'heure et par enfant depuis le 1er juin 2026 "
          "(avenant n° 10 à la convention IDCC 3239)."),
    ("p", "Indemnité d'entretien : 0,435 € par heure d'accueil, avec un plancher de 2,65 € par journée."),
    ("p", "Au-delà de 45 heures par semaine, les heures sont majorées : le taux se négocie au contrat."),
    ("p", "Congés payés : la méthode des 10 % et celle du maintien de salaire se comparent, et c'est la plus "
          "favorable qui s'applique. L'onglet Année calcule les deux. Le versement mensuel est interdit."),
    ("", ""),
    ("p", "Ces montants étaient à jour au 1er juin 2026. Vérifiez-les chaque année : ils sont revalorisés."),
]
r = 5
for typ, txt in lignes:
    c = ws.cell(row=r, column=2, value=txt)
    if typ == "titre":
        c.font = Font(name=F, size=11, bold=True, color=TERRA)
    elif typ == "p":
        c.font = norm; c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 28 if len(txt) > 100 else 15
    r += 1

# ----------------------------------------------------------------- Contrats
ws = wb.create_sheet("Contrats")
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 40
for col in COLS_ENF:
    ws.column_dimensions[col].width = 15
ws.column_dimensions["H"].width = 62
ws.merge_cells("B2:F2"); ws["B2"] = "Un enfant par colonne"; ws["B2"].font = h1
ws["B2"].fill = fond_titre; ws["B2"].alignment = Alignment(vertical="center")
ws.row_dimensions[2].height = 28

for i, col in enumerate(COLS_ENF, start=1):
    c = ws[f"{col}4"]; c.value = f"Enfant {i}"
    c.font = h2; c.fill = fond_entete; c.border = bord
    c.alignment = Alignment(horizontal="center")
ws["B4"] = "À REMPLIR"
ws["B4"].font = Font(name=F, size=10, bold=True, color="FFFFFF"); ws["B4"].fill = fond_entete


def champ(cle, libelle, defauts, fmt=None, aide=""):
    ligne = LIG[cle]
    ws.cell(row=ligne, column=2, value=libelle).font = gras
    for col, defaut in zip(COLS_ENF, defauts):
        c = ws[f"{col}{ligne}"]; c.value = defaut
        c.font = saisie; c.fill = fond_saisie; c.border = bord
        if fmt: c.number_format = fmt
    a = ws.cell(row=ligne, column=8, value=aide); a.font = petit
    a.alignment = Alignment(wrap_text=True, vertical="center")


vide = [None] * (NB_ENFANTS - 1)
champ("prenom", "Prénom de l'enfant", ["Lucie"] + vide, None,
      "Laissez vide les colonnes des enfants que vous n'accueillez pas.")
champ("debut", "Début du contrat", ["01/09/2026"] + vide, None, "Pour mémoire.")
champ("taux", "Taux horaire brut (€)", [4.5] + vide, EUR,
      "Ce que vous avez négocié, par heure et par enfant.")
champ("heuresSem", "Heures d'accueil par semaine", [45] + vide, HEU,
      "Les heures prévues au contrat, pas celles réellement faites.")
champ("semaines", "Semaines d'accueil par an", [47] + vide, HEU,
      "47 en année complète. Moins si la famille ne vous confie pas l'enfant toute l'année.")
champ("entretienH", "Indemnité d'entretien par heure (€)", [0.435] * NB_ENFANTS, EUR,
      "Montant conventionnel au 1er juin 2026.")
champ("plancher", "Plancher d'entretien par journée (€)", [2.65] * NB_ENFANTS, EUR,
      "En dessous, c'est ce plancher qui s'applique.")
champ("repas", "Indemnité par repas (€)", [4.0] + vide, EUR,
      "Aucun barème légal : c'est ce qui est écrit au contrat.")
champ("km", "Indemnité kilométrique (€/km)", [0.529] * NB_ENFANTS, EUR,
      "Barème fiscal. Exonérée si vous tenez une feuille de route.")
champ("semCongés", "Semaines de congés payés acquises", [5] + vide, HEU,
      "Cinq semaines pour une année complète de travail. Sert au calcul du maintien de salaire.")

ws["B15"] = "CALCULÉ — NE PAS MODIFIER"
ws["B15"].font = Font(name=F, size=10, bold=True, color="FFFFFF")
ws["B15"].fill = PatternFill("solid", fgColor=SAUGE)


def calc(cle, libelle, modele, fmt=None, aide=""):
    ligne = LIG[cle]
    ws.cell(row=ligne, column=2, value=libelle).font = gras
    for col in COLS_ENF:
        c = ws[f"{col}{ligne}"]; c.value = modele.format(c=col)
        c.font = Font(name=F, size=10, bold=True, color=MARINE)
        c.fill = fond_doux; c.border = bord
        if fmt: c.number_format = fmt
    a = ws.cell(row=ligne, column=8, value=aide); a.font = petit
    a.alignment = Alignment(wrap_text=True, vertical="center")


calc("heuresMens", "Heures mensualisées",
     "=IF({c}%d=\"\",\"\",ROUND({c}%d*{c}%d/12,2))" % (LIG["taux"], LIG["heuresSem"], LIG["semaines"]),
     HEU, "Heures par semaine × semaines par an ÷ 12. C'est la base payée chaque mois, même les mois creux.")
calc("brutMens", "Salaire mensualisé brut",
     "=IF({c}%d=\"\",\"\",ROUND({c}%d*{c}%d,2))" % (LIG["taux"], LIG["heuresMens"], LIG["taux"]),
     EUR, "Le montant dû tous les mois, indépendamment des heures réellement faites.")
calc("ctrlMin", "Contrôle du minimum conventionnel",
     "=IF({c}%d=\"\",\"\",IF({c}%d<4.2,\"ATTENTION : sous le minimum de 4,20 €\",\"Conforme\"))"
     % (LIG["taux"], LIG["taux"]),
     None, "Le minimum s'apprécie par heure et par enfant.")
calc("ctrlCMG", "Contrôle du plafond CMG",
     "=IF({c}%d=\"\",\"\",IF({c}%d>61.55,\"ATTENTION : au-delà du plafond, le CMG est perdu en entier\",\"Sous le plafond\"))"
     % (LIG["taux"], LIG["taux"]),
     None, "Cinq fois le SMIC horaire par jour et par enfant.")
for col in COLS_ENF:
    ws[f"{col}{LIG['ctrlMin']}"].number_format = "General"
    ws[f"{col}{LIG['ctrlCMG']}"].number_format = "General"
ws[f"C{LIG['prenom']}"].comment = Comment(
    "Exemple fourni : remplacez-le par vos propres valeurs.", "TiMat")

# ------------------------------------------------------------------- Heures
ws = wb.create_sheet("Heures")
entetes = ["Date", "Enfant", "Arrivée", "Départ", "Heures", "Repas", "Km", "Entretien (€)", "Clé mois"]
larg = [13, 9, 11, 11, 11, 9, 9, 15, 12]
ws.merge_cells("A1:I1"); ws["A1"] = "Une ligne par journée d'accueil"; ws["A1"].font = h1
ws["A1"].fill = fond_titre; ws["A1"].alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 26
ws["A2"] = ("Remplissez Date, Enfant (1 à 4), Arrivée, Départ, Repas et Km. "
            "Les colonnes grises se calculent seules.")
ws["A2"].font = petit
for i, (e, w) in enumerate(zip(entetes, larg), start=1):
    c = ws.cell(row=3, column=i, value=e)
    c.font = h2; c.fill = fond_entete; c.border = bord
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[3].height = 22

ent_h = f"Contrats!$C${LIG['entretienH']}:$F${LIG['entretienH']}"
pla_h = f"Contrats!$C${LIG['plancher']}:$F${LIG['plancher']}"
for r in range(PREMIERE, DERNIERE + 1):
    for col in (1, 2, 3, 4, 6, 7):
        c = ws.cell(row=r, column=col); c.font = saisie; c.fill = fond_saisie; c.border = bord
    ws.cell(row=r, column=1).number_format = DATE
    ws.cell(row=r, column=2).number_format = ENT
    ws.cell(row=r, column=3).number_format = HM
    ws.cell(row=r, column=4).number_format = HM
    d = ws.cell(row=r, column=5, value=f'=IF(OR(C{r}="",D{r}=""),"",ROUND((D{r}-C{r})*24,2))')
    d.number_format = HEU
    # Le plancher journalier prime quand la journée est courte : c'est la règle
    # conventionnelle, et l'erreur la plus fréquente des tableurs faits maison.
    e = ws.cell(row=r, column=8, value=(
        f'=IF(OR(E{r}="",B{r}=""),"",'
        f'MAX(INDEX({pla_h},B{r}),ROUND(E{r}*INDEX({ent_h},B{r}),2)))'))
    e.number_format = EUR
    # Clé numérique plutôt qu'un format de date : indépendante de la langue du tableur.
    ws.cell(row=r, column=9, value=f'=IF(A{r}="","",YEAR(A{r})*100+MONTH(A{r}))')
    for col in (5, 8, 9):
        c = ws.cell(row=r, column=col); c.font = norm; c.fill = fond_doux; c.border = bord
ws["A4"] = "01/09/2026"; ws["B4"] = 1; ws["C4"] = "07:30"; ws["D4"] = "18:00"
ws["E4"] = None; ws["F4"] = 1; ws["G4"] = 0
ws["A4"].comment = Comment("Ligne d'exemple : écrasez-la par votre première journée réelle.", "TiMat")
ws.freeze_panes = "A4"

H_HEU = f"Heures!$E${PREMIERE}:$E${DERNIERE}"
H_ENF = f"Heures!$B${PREMIERE}:$B${DERNIERE}"
H_REP = f"Heures!$F${PREMIERE}:$F${DERNIERE}"
H_KM = f"Heures!$G${PREMIERE}:$G${DERNIERE}"
H_ENT = f"Heures!$H${PREMIERE}:$H${DERNIERE}"
H_CLE = f"Heures!$I${PREMIERE}:$I${DERNIERE}"

# --------------------------------------------------------------------- Mois
ws = wb.create_sheet("Mois")
ws.merge_cells("A1:K1"); ws["A1"] = "Le détail mois par mois"; ws["A1"].font = h1
ws["A1"].fill = fond_titre; ws["A1"].alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 26
ws["A2"] = "Choisissez l'enfant, l'année et le premier mois. Le reste se calcule."
ws["A2"].font = petit
for cell, libelle, defaut in (("A3", "Enfant n°", 1), ("C3", "Année", 2026), ("E3", "Mois de départ (1 à 12)", 1)):
    ws[cell] = libelle; ws[cell].font = gras
    suivante = ws[cell].offset(column=1)
    suivante.value = defaut; suivante.font = saisie
    suivante.fill = fond_saisie; suivante.border = bord
ws["G3"] = "=IF(INDEX(Contrats!$C$5:$F$5,$B$3)=\"\",\"(colonne vide dans Contrats)\",INDEX(Contrats!$C$5:$F$5,$B$3))"
ws["G3"].font = Font(name=F, size=10, bold=True, color=TERRA)

cols = ["Mois", "Clé", "Heures faites", "Heures mensualisées", "Heures compl.", "Journées",
        "Brut mensualisé", "Brut compl.", "Entretien", "Repas", "Kilomètres"]
larg = [14, 10, 14, 19, 13, 11, 16, 14, 13, 12, 13]
for i, (e, w) in enumerate(zip(cols, larg), start=1):
    c = ws.cell(row=5, column=i, value=e)
    c.font = h2; c.fill = fond_entete; c.border = bord
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[5].height = 30


def idx(cle):
    """Le paramètre de l'enfant sélectionné en B3."""
    ligne = LIG[cle]
    return f"INDEX(Contrats!$C${ligne}:$F${ligne},$B$3)"


for n in range(12):
    r = 6 + n
    ws.cell(row=r, column=1, value=f'=DATE($D$3,$F$3+{n},1)').number_format = "MMMM YYYY"
    ws.cell(row=r, column=2, value=f'=YEAR(A{r})*100+MONTH(A{r})')
    ws.cell(row=r, column=3, value=f'=ROUND(SUMIFS({H_HEU},{H_CLE},B{r},{H_ENF},$B$3),2)').number_format = HEU
    ws.cell(row=r, column=4, value=f'=IFERROR({idx("heuresMens")},0)').number_format = HEU
    ws.cell(row=r, column=5, value=f'=ROUND(MAX(0,C{r}-D{r}),2)').number_format = HEU
    ws.cell(row=r, column=6, value=f'=COUNTIFS({H_CLE},B{r},{H_ENF},$B$3,{H_HEU},">0")').number_format = ENT
    # Le brut mensualisé reste dû les mois creux : il n'est mis à zéro que si
    # aucune heure n'a été saisie du tout pour le mois.
    ws.cell(row=r, column=7, value=f'=IF(C{r}=0,0,IFERROR({idx("brutMens")},0))').number_format = EUR
    ws.cell(row=r, column=8, value=f'=ROUND(E{r}*IFERROR({idx("taux")},0),2)').number_format = EUR
    ws.cell(row=r, column=9, value=f'=ROUND(SUMIFS({H_ENT},{H_CLE},B{r},{H_ENF},$B$3),2)').number_format = EUR
    ws.cell(row=r, column=10, value=(
        f'=ROUND(SUMIFS({H_REP},{H_CLE},B{r},{H_ENF},$B$3)*IFERROR({idx("repas")},0),2)')).number_format = EUR
    ws.cell(row=r, column=11, value=(
        f'=ROUND(SUMIFS({H_KM},{H_CLE},B{r},{H_ENF},$B$3)*IFERROR({idx("km")},0),2)')).number_format = EUR
    for col in range(1, 12):
        c = ws.cell(row=r, column=col); c.font = norm; c.border = bord
        if col in (7, 8): c.fill = fond_doux

r = 18
ws.cell(row=r, column=1, value="Total").font = Font(name=F, size=10, bold=True, color="FFFFFF")
ws.cell(row=r, column=1).fill = fond_titre
ws.cell(row=r, column=2).fill = fond_titre
for col in range(3, 12):
    L = get_column_letter(col)
    c = ws.cell(row=r, column=col, value=f"=SUM({L}6:{L}17)")
    c.font = Font(name=F, size=10, bold=True, color=MARINE); c.fill = fond_doux; c.border = bord
    c.number_format = ENT if col == 6 else (EUR if col >= 7 else HEU)

ws["A20"] = ("Le brut mensualisé est dû même les mois où l'enfant vient moins : c'est le principe de la "
             "mensualisation. Il n'est mis à zéro ici que si aucune heure n'a été saisie du tout pour le mois.")
ws["A20"].font = petit; ws.merge_cells("A20:K20")
ws["A21"] = ("Les heures complémentaires sont celles qui dépassent l'horaire mensualisé. Au-delà de 45 heures "
             "par semaine, un taux majoré peut s'appliquer : il se négocie au contrat et n'est pas calculé ici.")
ws["A21"].font = petit; ws.merge_cells("A21:K21")
ws.freeze_panes = "A6"

# -------------------------------------------------------------------- Année
ws = wb.create_sheet("Année")
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 38
for col in COLS_ENF:
    ws.column_dimensions[col].width = 15
ws.column_dimensions["G"].width = 15
ws.column_dimensions["H"].width = 58
ws.merge_cells("B2:G2"); ws["B2"] = "Les quatre contrats sur l'année"; ws["B2"].font = h1
ws["B2"].fill = fond_titre; ws["B2"].alignment = Alignment(vertical="center")
ws.row_dimensions[2].height = 28

for i, col in enumerate(COLS_ENF, start=1):
    c = ws[f"{col}4"]
    c.value = f'=IF(Contrats!{col}{LIG["prenom"]}="","Enfant {i}",Contrats!{col}{LIG["prenom"]})'
    c.font = h2; c.fill = fond_entete; c.border = bord
    c.alignment = Alignment(horizontal="center")
g = ws["G4"]; g.value = "Tous"
g.font = h2; g.fill = PatternFill("solid", fgColor=SAUGE); g.border = bord
g.alignment = Alignment(horizontal="center")

# Zone de calcul : le nombre de mois réellement travaillés par enfant. Il faut
# compter les mois distincts, ce qu'aucune fonction simple ne fait — d'où ce
# petit tableau, douze lignes par enfant, additionné ensuite.
DEB_AIDE = 40
ws.cell(row=DEB_AIDE - 1, column=2, value="Zone de calcul — ne pas modifier").font = petit
for n in range(12):
    r = DEB_AIDE + n
    ws.cell(row=r, column=2, value=f"=Mois!B{6 + n}").font = petit
    for i, col in enumerate(COLS_ENF, start=1):
        ws[f"{col}{r}"] = (f'=IF(COUNTIFS({H_CLE},$B{r},{H_ENF},{i},{H_HEU},">0")>0,1,0)')
        ws[f"{col}{r}"].font = petit
FIN_AIDE = DEB_AIDE + 11


def total(ligne, libelle, modele, aide, fmt=EUR, somme=True):
    ws.cell(row=ligne, column=2, value=libelle).font = gras
    for i, col in enumerate(COLS_ENF, start=1):
        c = ws[f"{col}{ligne}"]; c.value = modele.format(n=i, c=col)
        c.font = Font(name=F, size=10, bold=True, color=MARINE)
        c.fill = fond_doux; c.border = bord; c.number_format = fmt
    t = ws[f"G{ligne}"]
    t.value = f"=SUM(C{ligne}:F{ligne})" if somme else ""
    t.font = Font(name=F, size=10, bold=True, color=MARINE)
    t.fill = PatternFill("solid", fgColor="E8F1EF"); t.border = bord; t.number_format = fmt
    a = ws.cell(row=ligne, column=8, value=aide); a.font = petit
    a.alignment = Alignment(wrap_text=True, vertical="center")


total(6, "Heures réalisées", f'=ROUND(SUMIFS({H_HEU},{H_ENF},{{n}}),2)',
      "Somme des heures saisies dans l'onglet Heures.", HEU)
total(7, "Journées d'accueil", f'=COUNTIFS({H_ENF},{{n}},{H_HEU},">0")',
      "Utile pour vérifier l'indemnité d'entretien.", ENT)
total(8, "Mois travaillés", f"=SUM({{c}}{DEB_AIDE}:{{c}}{FIN_AIDE})",
      "Nombre de mois où au moins une heure a été saisie.", ENT)
total(9, "Salaire brut mensualisé",
      f'=ROUND({{c}}8*IFERROR(Contrats!{{c}}{LIG["brutMens"]},0),2)',
      "Le mensualisé, multiplié par le nombre de mois travaillés.")
total(10, "Heures complémentaires",
      f'=ROUND(MAX(0,{{c}}6-{{c}}8*IFERROR(Contrats!{{c}}{LIG["heuresMens"]},0))*IFERROR(Contrats!{{c}}{LIG["taux"]},0),2)',
      "Les heures au-delà de l'horaire mensualisé, sur l'ensemble de l'année.")
total(11, "Salaire brut total", "=ROUND({c}9+{c}10,2)",
      "C'est la base de l'indemnité de congés payés.")

ws["B13"] = "Congés payés : les deux méthodes"
ws["B13"].font = Font(name=F, size=11, bold=True, color=TERRA)
total(14, "Méthode 10 %", "=ROUND({c}11*0.1,2)",
      "Dix pour cent de la rémunération brute totale.")
total(15, "Méthode maintien de salaire",
      f'=ROUND(IFERROR(Contrats!{{c}}{LIG["semCongés"]},0)*IFERROR(Contrats!{{c}}{LIG["heuresSem"]},0)'
      f'*IFERROR(Contrats!{{c}}{LIG["taux"]},0),2)',
      "Ce que vous auriez gagné en travaillant les semaines de congés acquises.")
total(16, "Indemnité retenue", "=ROUND(MAX({c}14,{c}15),2)",
      "C'est la plus favorable des deux qui s'applique — jamais la moyenne, jamais celle qui arrange l'employeur.")

ws["B18"] = "Indemnités — non soumises à cotisations"
ws["B18"].font = Font(name=F, size=11, bold=True, color=TERRA)
total(19, "Indemnité d'entretien", f'=ROUND(SUMIFS({H_ENT},{H_ENF},{{n}}),2)',
      "Non soumise à cotisations, et non imposable.")
total(20, "Indemnités de repas",
      f'=ROUND(SUMIFS({H_REP},{H_ENF},{{n}})*IFERROR(Contrats!{{c}}{LIG["repas"]},0),2)',
      "Non soumises à cotisations dans la limite du barème.")
total(21, "Indemnités kilométriques",
      f'=ROUND(SUMIFS({H_KM},{H_ENF},{{n}})*IFERROR(Contrats!{{c}}{LIG["km"]},0),2)',
      "Exonérées si vous tenez une feuille de route mensuelle.")
total(22, "Total des indemnités", "=ROUND({c}19+{c}20+{c}21,2)",
      "Se déclarent à part du salaire sur Pajemploi.")
total(24, "Total perçu sur l'année", "=ROUND({c}11+{c}16+{c}22,2)",
      "Brut, congés payés et indemnités réunis. Le brut n'est pas le net : les cotisations sont calculées par Pajemploi.")

ws["B26"] = "À ne pas oublier"; ws["B26"].font = Font(name=F, size=11, bold=True, color=TERRA)
notes = [
    "En année complète, le salaire mensualisé est versé les douze mois, congés compris : la comparaison des deux "
    "méthodes sert surtout en année incomplète, et pour vérifier qu'on ne perd rien.",
    "Le versement mensuel des congés payés est interdit : ils se versent en une fois, ou au moment de la prise.",
    "L'indemnité d'entretien et les repas ne sont pas du salaire : ils ne se déclarent pas dans la même case.",
    "Pour la déclaration Pajemploi, la fenêtre court du 25 du mois au 5 du mois suivant.",
    "Ce classeur est une aide au calcul, pas un bulletin de salaire : c'est Pajemploi qui l'établit.",
]
for i, n in enumerate(notes):
    c = ws.cell(row=27 + i, column=2, value="• " + n); c.font = norm
    ws.merge_cells(start_row=27 + i, start_column=2, end_row=27 + i, end_column=8)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[27 + i].height = 26

ws["B34"] = "Kit TiMat · timat.app · repères à jour au 1er juin 2026, à vérifier chaque année."
ws["B34"].font = petit

for s in wb.worksheets:
    s.sheet_view.showGridLines = False

SORTIE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "documents")
os.makedirs(SORTIE, exist_ok=True)
wb.save(os.path.join(SORTIE, "kit-gestion-assmat.xlsx"))
print("[kit] kit-gestion-assmat.xlsx généré dans documents/")
