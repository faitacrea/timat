# -*- coding: utf-8 -*-
"""Contrôle des références du classeur, à défaut de pouvoir le recalculer.

LibreOffice ne répond pas dans l'environnement de développement, donc aucune
preuve que l'arithmétique tombe juste. Ce que ce script prouve en revanche,
c'est le défaut le plus probable d'un classeur écrit par programme : une
formule qui pointe une colonne ou une ligne à côté. Chaque référence est
résolue et confrontée à l'en-tête ou au libellé qu'elle est censée désigner.
"""
import os
import re
import sys
from openpyxl import load_workbook

RACINE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
wb = load_workbook(os.path.join(RACINE, "documents", "kit-gestion-assmat.xlsx"))
contrats, heures, mois, annee = wb["Contrats"], wb["Heures"], wb["Mois"], wb["Année"]
echecs = []


def attendu(condition, message):
    if not condition:
        echecs.append(message)


# 1. Chaque ligne de paramètre porte le libellé que les formules supposent.
libelles = {
    5: "Prénom de l'enfant", 7: "Taux horaire brut", 8: "Heures d'accueil par semaine",
    9: "Semaines d'accueil par an", 10: "Indemnité d'entretien par heure",
    11: "Plancher d'entretien par journée", 12: "Indemnité par repas",
    13: "Indemnité kilométrique", 14: "Semaines de congés payés acquises",
    16: "Heures mensualisées", 17: "Salaire mensualisé brut",
}
for ligne, debut in libelles.items():
    lu = str(contrats.cell(row=ligne, column=2).value or "")
    attendu(lu.startswith(debut), f"Contrats ligne {ligne} devait être « {debut} », trouvé « {lu} »")

# 2. Les quatre colonnes d'enfants existent, et dans cet ordre.
for i, col in enumerate("CDEF", start=1):
    lu = contrats[f"{col}4"].value
    attendu(lu == f"Enfant {i}", f"Contrats!{col}4 devait être « Enfant {i} », trouvé « {lu} »")

# 3. Les colonnes de l'onglet Heures, dans l'ordre supposé par les SUMIFS.
colonnes = {1: "Date", 2: "Enfant", 3: "Arrivée", 4: "Départ", 5: "Heures",
            6: "Repas", 7: "Km", 8: "Entretien (€)", 9: "Clé mois"}
for col, titre in colonnes.items():
    lu = heures.cell(row=3, column=col).value
    attendu(lu == titre, f"Heures colonne {col} devait être « {titre} », trouvé « {lu} »")

# 4. Les colonnes de l'onglet Mois.
colM = {1: "Mois", 2: "Clé", 3: "Heures faites", 4: "Heures mensualisées",
        5: "Heures compl.", 6: "Journées", 7: "Brut mensualisé", 8: "Brut compl.",
        9: "Entretien", 10: "Repas", 11: "Kilomètres"}
for col, titre in colM.items():
    lu = mois.cell(row=5, column=col).value
    attendu(lu == titre, f"Mois colonne {col} devait être « {titre} », trouvé « {lu} »")

# 5. Toute référence à une ligne de Contrats vise un paramètre réellement défini.
for ws in (heures, mois, annee):
    for rang in ws.iter_rows():
        for c in rang:
            if isinstance(c.value, str) and c.value.startswith("="):
                for n in re.findall(r"Contrats!\$?[C-F]\$?(\d+)", c.value):
                    attendu(int(n) in libelles,
                            f"{ws.title}!{c.coordinate} référence Contrats ligne {n}, qui n'est pas un paramètre")

# 6. Les plages SUMIFS couvrent exactement les lignes de saisie, sans déborder.
premiere = 4
derniere = max(r for r in range(4, heures.max_row + 1)
               if isinstance(heures.cell(row=r, column=9).value, str))
for ws in (mois, annee):
    for rang in ws.iter_rows():
        for c in rang:
            if isinstance(c.value, str) and "Heures!" in c.value:
                for debut, fin in re.findall(r"Heures!\$[A-I]\$(\d+):\$[A-I]\$(\d+)", c.value):
                    attendu(int(debut) == premiere and int(fin) == derniere,
                            f"{ws.title}!{c.coordinate} couvre {debut}-{fin}, attendu {premiere}-{derniere}")

# 7. L'onglet Année : les libellés attendus aux lignes que les formules chaînent.
attendus_annee = {
    6: "Heures réalisées", 7: "Journées d'accueil", 8: "Mois travaillés",
    9: "Salaire brut mensualisé", 10: "Heures complémentaires", 11: "Salaire brut total",
    14: "Méthode 10 %", 15: "Méthode maintien de salaire", 16: "Indemnité retenue",
    19: "Indemnité d'entretien", 20: "Indemnités de repas", 21: "Indemnités kilométriques",
    22: "Total des indemnités", 24: "Total perçu sur l'année",
}
for ligne, libelle in attendus_annee.items():
    lu = str(annee.cell(row=ligne, column=2).value or "")
    attendu(lu == libelle, f"Année ligne {ligne} devait être « {libelle} », trouvé « {lu} »")

# 8. La zone de calcul des mois travaillés couvre bien douze lignes, et la
#    ligne « Mois travaillés » les additionne toutes.
formule_mois = str(annee["C8"].value)
plage = re.search(r"SUM\(C(\d+):C(\d+)\)", formule_mois)
attendu(plage is not None, "Année!C8 devait additionner la zone de calcul")
if plage:
    debut, fin = int(plage.group(1)), int(plage.group(2))
    attendu(fin - debut == 11, f"la zone de calcul couvre {fin - debut + 1} lignes, attendu 12")
    for r in range(debut, fin + 1):
        attendu(str(annee.cell(row=r, column=2).value or "").startswith("=Mois!B"),
                f"Année!B{r} devait reprendre une clé de mois")

# 9. L'indemnité retenue compare bien les deux méthodes, dans les deux sens.
retenue = str(annee["C16"].value)
attendu("MAX(" in retenue and "C14" in retenue and "C15" in retenue,
        "Année!C16 doit retenir le maximum des deux méthodes de congés payés")

# 10. La ligne d'exemple doit être exploitable : de vraies valeurs de date et
#     d'heure, et surtout ses formules intactes. Écrire dans une cellule de
#     formule après coup l'efface sans bruit — c'est arrivé, et le classeur
#     livré ne calculait rien.
from datetime import date as _date, time as _time
attendu(isinstance(heures["A4"].value, _date), "Heures!A4 doit porter une vraie date, pas du texte")
attendu(isinstance(heures["C4"].value, _time), "Heures!C4 doit porter une vraie heure, pas du texte")
attendu(isinstance(heures["D4"].value, _time), "Heures!D4 doit porter une vraie heure, pas du texte")
for coord in ("E4", "H4", "I4"):
    v = heures[coord].value
    attendu(isinstance(v, str) and v.startswith("="),
            f"Heures!{coord} a perdu sa formule (valeur : {v!r})")

# 11. Le classeur doit demander un recalcul complet à l'ouverture : openpyxl
#     n'écrit aucune valeur en cache, et un lecteur qui se fie au cache
#     n'afficherait que des cellules vides.
attendu(getattr(wb.calculation, "fullCalcOnLoad", False) is True,
        "le classeur doit porter fullCalcOnLoad pour se recalculer à l'ouverture")

# 12. Aucune fonction récente qu'un tableur ancien pourrait ne pas évaluer.
#     La frontière de mot compte : SUMIFS et COUNTIFS contiennent « IFS( »
#     sans être la fonction IFS. Sans \b, le contrôle rejette des formules saines.
interdites = ("XLOOKUP", "XMATCH", "TEXTJOIN", "IFS", "SWITCH", "MAXIFS", "MINIFS",
              "FILTER", "UNIQUE", "SORT", "SEQUENCE")
motif = re.compile(r"\b(" + "|".join(interdites) + r")\s*\(")
for ws in wb.worksheets:
    for rang in ws.iter_rows():
        for c in rang:
            if isinstance(c.value, str) and c.value.startswith("="):
                trouve = motif.search(c.value.upper())
                attendu(trouve is None,
                        f"{ws.title}!{c.coordinate} utilise {trouve.group(1) if trouve else ''}, à éviter")

total = sum(1 for ws in wb.worksheets for rang in ws.iter_rows() for c in rang
            if isinstance(c.value, str) and c.value.startswith("="))
print(f"{total} formules contrôlées, {len(echecs)} anomalie(s)")
for e in echecs:
    print("  ✗ " + e)
sys.exit(1 if echecs else 0)
